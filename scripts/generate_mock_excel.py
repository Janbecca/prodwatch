import argparse
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADERS = [
    "project_id",
    "platform_code",
    "keyword",
    "raw_text",
    "publish_time",
    "platform_post_id",
    "like_count",
    "comment_count",
    "share_count",
]

PLATFORMS = ["weibo", "xhs", "douyin"]
KEYWORDS = ["萤石", "监控", "摄像头", "智能家居"]


def _random_publish_time() -> str:
    now = datetime.utcnow()
    delta = timedelta(seconds=random.randint(0, 24 * 60 * 60))
    return (now - delta).strftime("%Y-%m-%d %H:%M:%S")


def generate_rows(count: int):
    rows = []
    for i in range(count):
        platform = random.choice(PLATFORMS)
        keyword = random.choice(KEYWORDS)
        publish_time = _random_publish_time()
        platform_post_id = f"{platform}_{int(datetime.utcnow().timestamp())}_{i}"
        raw_text = f"{platform} 平台 关于 {keyword} 的样例内容 {i}"
        rows.append(
            [
                random.choice([1, 2]),
                platform,
                keyword,
                raw_text,
                publish_time,
                platform_post_id,
                random.randint(0, 200),
                random.randint(0, 50),
                random.randint(0, 30),
            ]
        )
    return rows


def write_excel(path: Path, rows):
    if path.exists():
        wb = load_workbook(path)
        if "post_raw_upload" in wb.sheetnames:
            ws = wb["post_raw_upload"]
            wb.remove(ws)
        ws = wb.create_sheet("post_raw_upload")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "post_raw_upload"

    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--count", type=int, default=20)
    parser.add_argument("--target", type=str, default="mock_upload.xlsx")
    args = parser.parse_args()

    rows = generate_rows(args.count)
    write_excel(Path(args.target), rows)


if __name__ == "__main__":
    main()
