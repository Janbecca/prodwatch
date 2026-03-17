import hashlib
import random
from datetime import datetime, timedelta

import openpyxl


DB_PATH = "backend/database/prodwatch_database.xlsx"


def _sheet_headers(ws):
    return [cell.value for cell in ws[1]]


def _max_numeric_id(ws, id_col_idx=1) -> int:
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        v = row[id_col_idx - 1]
        if isinstance(v, (int, float)) and v == v:
            try:
                max_id = max(max_id, int(v))
            except Exception:
                pass
    return max_id


def _ensure_sheet(wb, name: str, headers: list[str]):
    if name in wb.sheetnames:
        ws = wb[name]
        existing = _sheet_headers(ws)
        if existing and any(h not in existing for h in headers):
            for h in headers:
                if h not in existing:
                    ws.cell(row=1, column=len(existing) + 1, value=h)
                    existing.append(h)
        return ws
    ws = wb.create_sheet(name)
    ws.append(headers)
    return ws


def _sha256_hex(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def main():
    wb = openpyxl.load_workbook(DB_PATH)
    now = datetime.utcnow()

    # --- ensure required sheets/columns exist ---
    ws_run = wb["pipeline_run"]
    ws_raw = wb["post_raw"]
    ws_clean = wb["post_clean"]
    ws_spam = wb["spam_score"]
    ws_sent = wb["sentiment_result"]
    ws_kw = wb["monitor_keyword"]

    # these should already exist; still ensure columns are present (append missing at end)
    _ensure_sheet(
        wb,
        "pipeline_run",
        ["id", "project_id", "run_no", "trigger_type", "status", "start_time", "end_time", "params", "created_at"],
    )
    _ensure_sheet(
        wb,
        "post_raw",
        [
            "id",
            "pipeline_run_id",
            "project_id",
            "platform_id",
            "keyword_id",
            "content_type",
            "platform_post_id",
            "author_id",
            "publish_time",
            "raw_text",
            "like_count",
            "comment_count",
            "share_count",
        ],
    )
    _ensure_sheet(
        wb,
        "post_clean",
        ["id", "post_raw_id", "pipeline_run_id", "project_id", "clean_text", "text_hash", "is_valid", "invalid_reason"],
    )
    _ensure_sheet(
        wb,
        "spam_score",
        ["id", "post_clean_id", "project_id", "score_total", "label", "rule_hits"],
    )
    _ensure_sheet(
        wb,
        "sentiment_result",
        ["id", "post_clean_id", "project_id", "polarity", "confidence", "intensity", "emotions"],
    )
    _ensure_sheet(
        wb,
        "monitor_keyword",
        ["id", "project_id", "keyword", "keyword_type", "weight", "is_active", "created_at"],
    )

    run_headers = _sheet_headers(ws_run)
    raw_headers = _sheet_headers(ws_raw)
    clean_headers = _sheet_headers(ws_clean)
    spam_headers = _sheet_headers(ws_spam)
    sent_headers = _sheet_headers(ws_sent)
    kw_headers = _sheet_headers(ws_kw)

    # --- get existing dimensions ---
    project_ids = []
    if "monitor_project" in wb.sheetnames:
        ws_proj = wb["monitor_project"]
        for row in ws_proj.iter_rows(min_row=2, values_only=True):
            try:
                pid = int(row[0])
                if pid > 0:
                    project_ids.append(pid)
            except Exception:
                continue
    project_ids = sorted(set(project_ids)) or [1]

    platform_ids = []
    if "platform" in wb.sheetnames:
        ws_pl = wb["platform"]
        for row in ws_pl.iter_rows(min_row=2, values_only=True):
            try:
                pid = int(row[0])
                if pid > 0:
                    platform_ids.append(pid)
            except Exception:
                continue
    platform_ids = sorted(set(platform_ids)) or [1]

    # --- create a new pipeline run (so /posts?mode=latest_run shows these) ---
    run_id = int(now.timestamp() * 1000)
    run_no = now.strftime("%Y%m%d%H%M%S")
    run_row = {h: None for h in run_headers}
    run_row.update(
        {
            "id": run_id,
            "project_id": project_ids[0],
            "run_no": run_no,
            "trigger_type": "seed_fake_posts",
            "status": "finished",
            "start_time": now,
            "end_time": now,
            "params": None,
            "created_at": now,
        }
    )
    ws_run.append([run_row.get(h) for h in run_headers])

    # --- seed some keywords (re-used) ---
    existing_kw = set()
    existing_kw_map = {}  # (project_id, keyword) -> id
    for row in ws_kw.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        try:
            kid = int(row[0])
            pid = int(row[1]) if isinstance(row[1], (int, float)) and row[1] == row[1] else None
            kw = row[2]
            if pid and isinstance(kw, str) and kw.strip():
                key = (pid, kw.strip())
                existing_kw.add(key)
                existing_kw_map[key] = kid
        except Exception:
            continue

    next_kw_id = _max_numeric_id(ws_kw) + 1
    keywords = ["摄像头", "门锁", "夜视", "画质", "续航", "售后", "价格", "安装", "隐私", "功能"]
    for pid in project_ids[:5]:
        for kw in random.sample(keywords, k=min(4, len(keywords))):
            key = (pid, kw)
            if key in existing_kw:
                continue
            row = {h: None for h in kw_headers}
            row.update(
                {
                    "id": next_kw_id,
                    "project_id": pid,
                    "keyword": kw,
                    "keyword_type": None,
                    "weight": None,
                    "is_active": 1,
                    "created_at": now,
                }
            )
            ws_kw.append([row.get(h) for h in kw_headers])
            existing_kw.add(key)
            existing_kw_map[key] = next_kw_id
            next_kw_id += 1

    # --- seed posts ---
    next_raw_id = _max_numeric_id(ws_raw) + 1
    next_clean_id = _max_numeric_id(ws_clean) + 1
    next_spam_id = _max_numeric_id(ws_spam) + 1
    next_sent_id = _max_numeric_id(ws_sent) + 1

    # de-dup by platform_post_id
    existing_platform_post_ids = set()
    try:
        ppi_idx = raw_headers.index("platform_post_id")
        for row in ws_raw.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= ppi_idx:
                continue
            v = row[ppi_idx]
            if isinstance(v, str) and v.strip():
                existing_platform_post_ids.add(v.strip())
    except Exception:
        pass

    phrases_pos = ["画质不错", "很清晰", "安装方便", "性价比高", "值得推荐", "夜视效果好"]
    phrases_neu = ["刚入手", "在对比", "有点纠结", "准备装在家门口", "看评测中"]
    phrases_neg = ["卡顿严重", "经常掉线", "售后太差", "画面模糊", "发热厉害", "隐私担忧"]
    spam_tokens = ["加微信VX", "免费领取", "返现", "http://example.com"]

    n_posts = 90
    for i in range(n_posts):
        project_id = random.choice(project_ids)
        platform_id = random.choice(platform_ids)
        day = now.date() - timedelta(days=random.randint(0, 13))
        publish_time = datetime(day.year, day.month, day.day, random.randint(0, 23), random.randint(0, 59), random.randint(0, 59))

        pol = random.choices(["positive", "neutral", "negative"], weights=[0.35, 0.35, 0.30], k=1)[0]
        if pol == "positive":
            text = random.choice(phrases_pos)
        elif pol == "negative":
            text = random.choice(phrases_neg)
        else:
            text = random.choice(phrases_neu)

        # occasionally add spammy content
        is_spam = random.random() < 0.12
        if is_spam:
            text = f"{text} {random.choice(spam_tokens)} !!!"

        keyword = random.choice(keywords)
        keyword_id = existing_kw_map.get((project_id, keyword))
        if keyword_id is None:
            keyword_id = existing_kw_map.get((project_ids[0], keyword))

        platform_post_id = f"seed_{run_id}_{platform_id}_{next_raw_id}"
        if platform_post_id in existing_platform_post_ids:
            continue

        raw_row = {h: None for h in raw_headers}
        raw_row.update(
            {
                "id": next_raw_id,
                "pipeline_run_id": run_id,
                "project_id": project_id,
                "platform_id": platform_id,
                "keyword_id": keyword_id,
                "content_type": "post",
                "platform_post_id": platform_post_id,
                "author_id": None,
                "publish_time": publish_time,
                "raw_text": f"【{keyword}】{text}",
                "like_count": random.randint(0, 500),
                "comment_count": random.randint(0, 120),
                "share_count": random.randint(0, 80),
            }
        )
        ws_raw.append([raw_row.get(h) for h in raw_headers])
        existing_platform_post_ids.add(platform_post_id)

        # post_clean (valid unless empty)
        clean_text = str(raw_row["raw_text"] or "").strip()
        is_valid = 1 if clean_text else 0
        clean_row = {h: None for h in clean_headers}
        clean_row.update(
            {
                "id": next_clean_id,
                "post_raw_id": next_raw_id,
                "pipeline_run_id": run_id,
                "project_id": project_id,
                "clean_text": clean_text,
                "text_hash": _sha256_hex(clean_text) if clean_text else None,
                "is_valid": is_valid,
                "invalid_reason": None if is_valid == 1 else "empty",
            }
        )
        ws_clean.append([clean_row.get(h) for h in clean_headers])

        # spam_score
        spam_score = round(random.uniform(0.05, 0.25), 4)
        label = "normal"
        if is_spam:
            spam_score = round(random.uniform(0.65, 0.95), 4)
            label = "spam"
        elif spam_score >= 0.3:
            label = "suspect"
        spam_row = {h: None for h in spam_headers}
        spam_row.update(
            {
                "id": next_spam_id,
                "post_clean_id": next_clean_id,
                "project_id": project_id,
                "score_total": spam_score,
                "label": label,
                "rule_hits": "{}",
            }
        )
        ws_spam.append([spam_row.get(h) for h in spam_headers])

        # sentiment_result
        intensity = round(random.uniform(0.25, 0.9), 2)
        conf = round(random.uniform(0.55, 0.95), 2)
        if is_spam:
            pol_out = "neutral"
        else:
            pol_out = pol
        sent_row = {h: None for h in sent_headers}
        sent_row.update(
            {
                "id": next_sent_id,
                "post_clean_id": next_clean_id,
                "project_id": project_id,
                "polarity": pol_out,
                "confidence": conf,
                "intensity": intensity,
                "emotions": "{}",
            }
        )
        ws_sent.append([sent_row.get(h) for h in sent_headers])

        next_raw_id += 1
        next_clean_id += 1
        next_spam_id += 1
        next_sent_id += 1

    wb.save(DB_PATH)
    print("seeded posts:", DB_PATH, "run_id:", run_id)


if __name__ == "__main__":
    main()
