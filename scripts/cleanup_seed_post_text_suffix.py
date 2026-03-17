import re
from datetime import datetime

import openpyxl


DB_PATH = "backend/database/prodwatch_database.xlsx"


SUFFIX_RE = re.compile(r"（项目\d+\/平台\d+）$")


def _sheet_headers(ws):
    return [cell.value for cell in ws[1]]


def main() -> None:
    wb = openpyxl.load_workbook(DB_PATH)
    if "post_raw" not in wb.sheetnames:
        print("no post_raw sheet")
        return

    ws_raw = wb["post_raw"]
    headers = _sheet_headers(ws_raw)
    if "raw_text" not in headers:
        print("post_raw has no raw_text column")
        return
    raw_text_col = headers.index("raw_text") + 1

    cleaned = 0
    touched_raw_ids: set[int] = set()

    for r in ws_raw.iter_rows(min_row=2, values_only=False):
        cell = r[raw_text_col - 1]
        v = cell.value
        if not isinstance(v, str):
            continue
        nv = SUFFIX_RE.sub("", v).rstrip()
        if nv != v:
            cell.value = nv
            cleaned += 1
            # also record id for updating post_clean
            try:
                raw_id = r[0].value
                if isinstance(raw_id, (int, float)) and raw_id == raw_id:
                    touched_raw_ids.add(int(raw_id))
            except Exception:
                pass

    # Keep post_clean.clean_text in sync for the touched raw posts
    if cleaned and "post_clean" in wb.sheetnames and touched_raw_ids:
        ws_clean = wb["post_clean"]
        ch = _sheet_headers(ws_clean)
        if "post_raw_id" in ch and "clean_text" in ch:
            rid_col = ch.index("post_raw_id") + 1
            text_col = ch.index("clean_text") + 1
            for r in ws_clean.iter_rows(min_row=2, values_only=False):
                rid = r[rid_col - 1].value
                if isinstance(rid, (int, float)) and rid == rid:
                    rid_i = int(rid)
                else:
                    continue
                if rid_i not in touched_raw_ids:
                    continue
                v = r[text_col - 1].value
                if not isinstance(v, str):
                    continue
                nv = SUFFIX_RE.sub("", v).rstrip()
                if nv != v:
                    r[text_col - 1].value = nv

    wb.save(DB_PATH)
    print("cleaned", cleaned, "rows at", datetime.utcnow().isoformat())


if __name__ == "__main__":
    main()

