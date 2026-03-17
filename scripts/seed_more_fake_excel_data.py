import random
from datetime import datetime, timedelta

import openpyxl


DB_PATH = "backend/database/prodwatch_database.xlsx"


def _sheet_headers(ws):
    return [cell.value for cell in ws[1]]


def _max_numeric_id(ws, id_col_idx=1):
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        v = row[id_col_idx - 1]
        if isinstance(v, (int, float)) and v == v:
            try:
                max_id = max(max_id, int(v))
            except Exception:
                pass
    return max_id


def _existing_values(ws, col_idx):
    out = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col_idx:
            continue
        v = row[col_idx]
        if isinstance(v, str):
            out.add(v.strip())
    return out


def _ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        # Ensure header row contains required columns (append missing at end).
        existing = _sheet_headers(ws)
        if existing and any(h not in existing for h in headers):
            for h in headers:
                if h not in existing:
                    ws.cell(row=1, column=len(existing) + 1, value=h)
                    existing.append(h)
        return ws
    ws = wb.create_sheet(name)
    ws.append(headers)
    return ws


def main():
    wb = openpyxl.load_workbook(DB_PATH)
    now = datetime.utcnow()

    # --- brand ---
    ws_brand = wb["brand"]
    brand_headers = _sheet_headers(ws_brand)
    # expected: id, name, industry, created_at
    existing_brand_names = _existing_values(ws_brand, 1)
    next_brand_id = _max_numeric_id(ws_brand) + 1

    new_brands = [
        ("TP-Link", "智能安防"),
        ("乔安", "智能安防"),
        ("360", "智能硬件"),
        ("华为智选", "智能硬件"),
    ]
    for name, industry in new_brands:
        if name in existing_brand_names:
            continue
        ws_brand.append([next_brand_id, name, industry, now])
        existing_brand_names.add(name)
        next_brand_id += 1

    # --- monitor_project ---
    ws_proj = wb["monitor_project"]
    proj_headers = _sheet_headers(ws_proj)
    # ensure new columns exist for our config UI
    required_proj_cols = [
        "id",
        "brand_id",
        "name",
        "description",
        "is_active",
        "created_at",
        "updated_at",
        "product_category",
    ]
    if proj_headers and any(h not in proj_headers for h in required_proj_cols):
        for h in required_proj_cols:
            if h not in proj_headers:
                ws_proj.cell(row=1, column=len(proj_headers) + 1, value=h)
                proj_headers.append(h)

    existing_proj_names = _existing_values(ws_proj, 2)
    next_proj_id = _max_numeric_id(ws_proj) + 1

    # Build a brand name -> id mapping
    brand_name_to_id = {}
    for row in ws_brand.iter_rows(min_row=2, values_only=True):
        if not row or not isinstance(row[0], (int, float)):
            continue
        bid = int(row[0])
        bname = row[1]
        if isinstance(bname, str) and bname.strip():
            brand_name_to_id[bname.strip()] = bid

    def bid(name: str) -> int:
        return int(brand_name_to_id[name])

    # projects are "monitoring topics" that can include multiple brands
    project_defs = [
        {
            "name": "摄像头舆情监控（演示）",
            "category": "摄像头",
            "brands": ["大华", "小米", "海康威视"],
            "desc": "演示项目：监控摄像头品类下的竞品舆情（多品牌）。",
        },
        {
            "name": "门锁舆情监控（演示）",
            "category": "智能门锁",
            "brands": ["小米", "华为智选"],
            "desc": "演示项目：监控智能门锁品类舆情（多品牌）。",
        },
        {
            "name": "路由器舆情监控（演示）",
            "category": "路由器",
            "brands": ["TP-Link", "小米", "360"],
            "desc": "演示项目：监控路由器品类舆情（多品牌）。",
        },
    ]

    seeded_project_ids = []
    for d in project_defs:
        if d["name"] in existing_proj_names:
            continue
        brands = d["brands"]
        legacy_brand_id = bid(brands[0]) if brands else None
        # Append values aligned to headers we expect.
        row = {h: None for h in proj_headers}
        row.update(
            {
                "id": next_proj_id,
                "brand_id": legacy_brand_id,
                "name": d["name"],
                "description": d["desc"],
                "is_active": 1,
                "created_at": now,
                "updated_at": now,
                "product_category": d["category"],
            }
        )
        ws_proj.append([row.get(h) for h in proj_headers])
        existing_proj_names.add(d["name"])
        seeded_project_ids.append(next_proj_id)
        next_proj_id += 1

    # --- monitor_project_brand (project<->brand join) ---
    ws_join = _ensure_sheet(
        wb,
        "monitor_project_brand",
        headers=["id", "project_id", "brand_id", "created_at"],
    )
    join_headers = _sheet_headers(ws_join)
    next_join_id = _max_numeric_id(ws_join) + 1
    existing_pairs = set()
    for row in ws_join.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        try:
            pid = int(row[1])
            bid_ = int(row[2])
            existing_pairs.add((pid, bid_))
        except Exception:
            continue

    # Fill join rows for both existing single-brand projects and newly created multi-brand projects.
    proj_name_to_id = {}
    for row in ws_proj.iter_rows(min_row=2, values_only=True):
        try:
            pid = int(row[0])
        except Exception:
            continue
        name = row[2]
        if isinstance(name, str) and name.strip():
            proj_name_to_id[name.strip()] = pid

    for d in project_defs:
        pid = proj_name_to_id.get(d["name"])
        if not pid:
            continue
        for bname in d["brands"]:
            bid_ = bid(bname)
            if (pid, bid_) in existing_pairs:
                continue
            row = {
                "id": next_join_id,
                "project_id": pid,
                "brand_id": bid_,
                "created_at": now,
            }
            ws_join.append([row.get(h) for h in join_headers])
            existing_pairs.add((pid, bid_))
            next_join_id += 1

    # --- daily_metric / sentiment_result: add some extra rows for the seeded projects to make dashboard more lively ---
    if seeded_project_ids and "daily_metric" in wb.sheetnames:
        ws_dm = wb["daily_metric"]
        dm_headers = _sheet_headers(ws_dm)
        next_dm_id = _max_numeric_id(ws_dm) + 1
        platform_ids = [1, 2, 3]
        start = datetime.utcnow().date() - timedelta(days=13)
        existing_keys = set()
        for row in ws_dm.iter_rows(min_row=2, values_only=True):
            if not row or not isinstance(row[0], (int, float)) or len(row) < 4:
                continue
            try:
                pid = int(row[1])
                pl = int(row[2])
                md = row[3]
                if isinstance(md, datetime):
                    md_key = md.date().isoformat()
                else:
                    md_key = str(md)
                existing_keys.add((pid, pl, md_key))
            except Exception:
                continue

        for d in range(14):
            day = start + timedelta(days=d)
            metric_date = datetime(day.year, day.month, day.day)
            md_key = day.isoformat()
            for project_id in seeded_project_ids:
                base = 45 + (project_id % 5) * 12
                for platform_id in platform_ids:
                    if (project_id, platform_id, md_key) in existing_keys:
                        continue
                    total = max(8, int(random.gauss(base, 10)))
                    valid = max(0, total - random.randint(0, max(1, int(total * 0.18))))
                    spam = max(0, int(valid * random.uniform(0.02, 0.20)))
                    remain = max(valid - spam, 0)
                    neg = max(0, int(remain * random.uniform(0.18, 0.50)))
                    pos = max(0, int(remain * random.uniform(0.12, 0.35)))
                    neu = max(remain - neg - pos, 0)
                    ws_dm.append([next_dm_id, project_id, platform_id, metric_date, total, valid, spam, pos, neu, neg])
                    next_dm_id += 1
                    existing_keys.add((project_id, platform_id, md_key))

    if seeded_project_ids and "sentiment_result" in wb.sheetnames:
        ws_sr = wb["sentiment_result"]
        next_sid = _max_numeric_id(ws_sr) + 1
        for project_id in seeded_project_ids:
            for _ in range(15):
                intensity = round(random.uniform(0.20, 0.90), 2)
                polarity = random.choice(["positive", "neutral", "negative"])
                conf = round(random.uniform(0.55, 0.95), 2)
                # columns: id, post_clean_id, project_id, polarity, confidence, intensity, emotions
                ws_sr.append([next_sid, next_sid, project_id, polarity, conf, intensity, "{}"])
                next_sid += 1

    wb.save(DB_PATH)
    print("seeded more:", DB_PATH)


if __name__ == "__main__":
    main()

