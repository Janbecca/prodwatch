import random
from datetime import datetime, timedelta

import openpyxl


DB_PATH = "backend/database/prodwatch_database.xlsx"


def _sheet_headers(ws):
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    return headers


def _max_numeric_id(ws, id_col_idx=1):
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[id_col_idx - 1]
        if isinstance(v, (int, float)) and v == v:
            try:
                max_id = max(max_id, int(v))
            except Exception:
                pass
    return max_id


def _ensure_row(ws, key_idx, key_value, row_values):
    for r in ws.iter_rows(min_row=2, values_only=False):
        if r[key_idx].value == key_value:
            return False
    ws.append(row_values)
    return True


def main():
    wb = openpyxl.load_workbook(DB_PATH)

    # --- platform ---
    ws = wb["platform"]
    headers = _sheet_headers(ws)
    # expected: id, code, name, created_at
    now = datetime.utcnow()
    existing_codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[1]
        if isinstance(code, str):
            existing_codes.add(code.strip())
    if "xhs" not in existing_codes:
        ws.append([2, "xhs", "小红书", now])
    if "douyin" not in existing_codes:
        ws.append([3, "douyin", "抖音", now])

    # --- brand ---
    ws_brand = wb["brand"]
    existing_brand_names = set()
    for row in ws_brand.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if isinstance(name, str):
            existing_brand_names.add(name.strip())
    brand_rows = [
        (2, "海康威视", "智能安防", now),
        (3, "大华", "智能安防", now),
        (4, "小米", "智能硬件", now),
    ]
    for r in brand_rows:
        if r[1] not in existing_brand_names:
            ws_brand.append(list(r))

    # --- monitor_project (competitors) ---
    ws_proj = wb["monitor_project"]
    existing_proj_names = set()
    for row in ws_proj.iter_rows(min_row=2, values_only=True):
        name = row[2]
        if isinstance(name, str):
            existing_proj_names.add(name.strip())
    proj_rows = [
        (2, 2, "海康威视摄像头竞品监控", "示例：监控海康竞品舆情", 1, now),
        (3, 3, "大华摄像头竞品监控", "示例：监控大华竞品舆情", 1, now),
        (4, 4, "小米摄像头竞品监控", "示例：监控小米竞品舆情", 1, now),
    ]
    for r in proj_rows:
        if r[2] not in existing_proj_names:
            ws_proj.append(list(r))

    # --- daily_metric (seed 30 days) ---
    ws_dm = wb["daily_metric"]
    dm_headers = _sheet_headers(ws_dm)
    # expected: id, project_id, platform_id, metric_date, total_posts, valid_posts, spam_posts, pos_posts, neu_posts, neg_posts
    next_id = _max_numeric_id(ws_dm) + 1
    project_ids = [1, 2, 3, 4]
    platform_ids = [1, 2, 3]
    start = datetime.utcnow().date() - timedelta(days=29)

    existing_keys = set()
    for row in ws_dm.iter_rows(min_row=2, values_only=True):
        if not row or not isinstance(row[0], (int, float)):
            continue
        try:
            project_id = int(row[1])
            platform_id = int(row[2])
            md = row[3]
            if isinstance(md, datetime):
                md_key = md.date().isoformat()
            else:
                md_key = str(md)
            existing_keys.add((project_id, platform_id, md_key))
        except Exception:
            continue

    for d in range(30):
        day = start + timedelta(days=d)
        metric_date = datetime(day.year, day.month, day.day)
        md_key = day.isoformat()
        for project_id in project_ids:
            base = 60 + project_id * 15
            for platform_id in platform_ids:
                if (project_id, platform_id, md_key) in existing_keys:
                    continue

                total = max(10, int(random.gauss(base, 18)))
                valid = max(0, total - random.randint(0, int(total * 0.15)))
                spam = max(0, int(valid * random.uniform(0.02, 0.15)))
                remain = max(valid - spam, 0)
                # Make one competitor have a noticeable negative spike on the last day to demonstrate alerts.
                if project_id == 2 and d == 29:
                    neg = max(0, int(remain * random.uniform(0.55, 0.75)))
                    pos = max(0, int(remain * random.uniform(0.05, 0.15)))
                elif project_id == 2 and d == 28:
                    neg = max(0, int(remain * random.uniform(0.10, 0.18)))
                    pos = max(0, int(remain * random.uniform(0.35, 0.55)))
                else:
                    neg = max(0, int(remain * random.uniform(0.18, 0.45)))
                    pos = max(0, int(remain * random.uniform(0.15, 0.40)))
                neu = max(remain - neg - pos, 0)
                ws_dm.append([next_id, project_id, platform_id, metric_date, total, valid, spam, pos, neu, neg])
                next_id += 1
                existing_keys.add((project_id, platform_id, md_key))

    # Force a visible alert spike by rewriting the latest two days for project_id=2
    # across all platforms (if those rows exist).
    def _to_datestr(v):
        if isinstance(v, datetime):
            return v.date().isoformat()
        try:
            return pd.to_datetime(v).date().isoformat()  # type: ignore
        except Exception:
            return str(v)

    # find latest two dates present for project 2
    import pandas as pd

    dates = []
    for row in ws_dm.iter_rows(min_row=2, values_only=True):
        if not row or not isinstance(row[0], (int, float)):
            continue
        try:
            if int(row[1]) == 2:
                md = row[3]
                if isinstance(md, datetime):
                    dates.append(md.date())
        except Exception:
            continue
    dates = sorted(set(dates))
    if len(dates) >= 2:
        d2 = dates[-2]
        d1 = dates[-1]

        def apply_values(target_date, total, valid, spam, pos, neu, neg):
            for r in ws_dm.iter_rows(min_row=2, values_only=False):
                try:
                    if not isinstance(r[0].value, (int, float)):
                        continue
                    if int(r[1].value) != 2:
                        continue
                    md = r[3].value
                    if not isinstance(md, datetime) or md.date() != target_date:
                        continue
                    # columns: id, project_id, platform_id, metric_date, total_posts, valid_posts, spam_posts, pos_posts, neu_posts, neg_posts
                    r[4].value = int(total)
                    r[5].value = int(valid)
                    r[6].value = int(spam)
                    r[7].value = int(pos)
                    r[8].value = int(neu)
                    r[9].value = int(neg)
                except Exception:
                    continue

        # previous day: lower negative
        apply_values(d2, total=120, valid=110, spam=5, pos=50, neu=45, neg=15)
        # latest day: higher negative spike
        apply_values(d1, total=130, valid=120, spam=15, pos=10, neu=30, neg=80)

    # --- sentiment_result (seed intensity per project) ---
    ws_sr = wb["sentiment_result"]
    sr_headers = _sheet_headers(ws_sr)
    next_sid = _max_numeric_id(ws_sr) + 1
    existing_counts = {pid: 0 for pid in project_ids}
    for row in ws_sr.iter_rows(min_row=2, values_only=True):
        try:
            if isinstance(row[0], (int, float)):
                pid = int(row[2])
                if pid in existing_counts:
                    existing_counts[pid] += 1
        except Exception:
            continue
    for project_id in project_ids:
        needed = max(0, 20 - existing_counts.get(project_id, 0))
        for _ in range(needed):
            intensity = round(random.uniform(0.25, 0.85), 2)
            polarity = random.choice(["positive", "neutral", "negative"])
            conf = round(random.uniform(0.55, 0.95), 2)
            ws_sr.append([next_sid, next_sid, project_id, polarity, conf, intensity, "{}"])
            next_sid += 1

    wb.save(DB_PATH)
    print("seeded:", DB_PATH)


if __name__ == "__main__":
    main()
